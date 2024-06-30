import time

import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as ec
from selenium.webdriver.support.wait import WebDriverWait
import openpyxl

file_path = "C:\\Users\\pradh\\downloads\\download.xlsx"
fruit = "Apple"
new_value = "700"


def update_execl(file_path_name, col_name, fruit_name, new_value_name):
    book = openpyxl.load_workbook(file_path_name)
    sheet = book.active
    dict1 = {}

    for i in range(1, sheet.max_column + 1):
        if sheet.cell(row=1, column=i).value == col_name:
            dict1["col"] = i

    for x in range(1, sheet.max_row + 1):
        for j in range(1, sheet.max_column + 1):
            if sheet.cell(row=x, column=j).value == fruit_name:
                dict1["row"] = x

    sheet.cell(row=dict1["row"], column=dict1["col"]).value = new_value_name
    book.save(file_path)


driver = webdriver.Chrome()
driver.maximize_window()
driver.implicitly_wait(5)
driver.get("https://rahulshettyacademy.com/upload-download-test/")
driver.find_element(By.ID, "downloadButton").click()
time.sleep(5)
update_execl(file_path, "price", fruit, new_value)

driver.find_element(By.XPATH, "//input[@type='file']").send_keys(file_path)

wait = WebDriverWait(driver, 10)
successful_text = wait.until(
    ec.visibility_of_element_located((By.CSS_SELECTOR, ".Toastify__toast-body :nth-child(2)"))).text

Price_column = driver.find_element(By.XPATH, "//div[text()='Price']").get_attribute("data-column-id")

actual_price = driver.find_element(By.XPATH,
                                   "//div[text()='" + fruit + "']/parent::div/parent::div/div[@id='cell-" + Price_column + "-undefined']")

print(successful_text)
print(actual_price.text)

assert actual_price.text == new_value
time.sleep(10)
