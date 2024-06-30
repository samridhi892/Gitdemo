import openpyxl


class HomePageData:
    test_HomePage_data = [{"first name": "Samridhi", "email": "sp@gmail.com", "gender": "Female"},
                          {"first name": "Vivek", "email": "vp@gmail.com", "gender": "Male"}]

    @staticmethod
    def getTestData(test_case_name):

        dic1 = {}
        book = openpyxl.load_workbook("C:\\The Folder\\Samridhi\\Selenium files\\test_worksheet.xlsx")
        sheet = book.active
        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == test_case_name:
                for j in range(2, sheet.max_column + 1):
                    dic1[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

        return [dic1]

