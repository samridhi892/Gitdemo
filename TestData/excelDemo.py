import openpyxl

book = openpyxl.load_workbook("C:\\The Folder\\Samridhi\\Selenium files\\test_worksheet.xlsx")
sheet = book.active
dic1 = {}
cell = sheet.cell(row=1, column=1)
print(cell.value)
# sheet.cell(row=1, column=3).value = "Samridhi"
print(sheet.cell(row=1, column=3).value)

for i in range(1, sheet.max_row + 1):
    if sheet.cell(row=i, column=1).value == "test case2":
        for j in range(2, sheet.max_column + 1):
            dic1[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

print(dic1)
