import openpyxl


def update_execl(file_path, col_name, fruit, new_value):
    book = openpyxl.load_workbook(file_path)
    sheet = book.active
    dict1 = {}

    for x in range(1, sheet.max_row + 1):
        for j in range(1, sheet.max_column + 1):
            if sheet.cell(row=x, column=j).value == fruit:
                dict1["row"] = x

    for i in range(1, sheet.max_row + 1):
        if sheet.cell(row=1, column=i) == col_name:
            dict1["col"] = i

    sheet.cell(row=dict1["row"], column=dict1["col"]).value = new_value
    book.save(file_path)

